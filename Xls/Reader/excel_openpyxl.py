from openpyxl import load_workbook
import json
import sys
import os
import argparse
import datetime


def run(argv):
    parser = argparse.ArgumentParser()
    parser.add_argument('--size')
    parser.add_argument('--start')
    parser.add_argument('--action')
    parser.add_argument('--file')
    parser.add_argument('--max-empty-rows', dest="max_empty_rows")
    parser.add_argument('--with-empty-rows')
    args = parser.parse_args()
    with_empty_rows = bool(args.with_empty_rows)

    if not os.path.isfile(args.file):
        print("File does not exist")
        sys.exit(1)

    workbook = load_workbook(args.file, read_only=True, data_only=True)
    sheet = workbook.active

    if args.action == "count":
        max_empty_rows = int(args.max_empty_rows)
        sheet.max_row = None
        total_rows_count = 0
        empty_rows_count = 0
        for row in sheet.iter_rows(row_offset=0):
            row_empty = True
            for cell in row:
                if cell.value is not None:
                    row_empty = False
                    empty_rows_count = 0
                    break

            if not row_empty or with_empty_rows:
                total_rows_count += 1

            if row_empty:
                empty_rows_count += 1

            if max_empty_rows < empty_rows_count:
                if with_empty_rows:
                    total_rows_count = total_rows_count - empty_rows_count  # cut off trailing empty rows
                break

        print(total_rows_count)

    elif args.action == "read":
        rows = []
        # sheet.max_row = sheet.max_column = None
        row_process_number = 0
        for row in sheet.iter_rows(row_offset=int(args.start) - 1):
            row_process_number += 1
            current_row_read = []
            row_empty = True
            for cell in row:
                value = cell.value
                if value is not None:
                    row_empty = False

                if isinstance(value, (datetime.datetime, datetime.time)):
                    current_row_read.append(value.isoformat())
                elif isinstance(value, (str, unicode)):
                    current_row_read.append(value.encode('utf-8'))
                else:
                    current_row_read.append(value)

            if not row_empty or with_empty_rows:
                rows.append(current_row_read)

            if row_process_number >= int(args.size):
                break

        print(json.dumps(rows))

    else:
        print("Unknown command")
        sys.exit(1)


if __name__ == "__main__":
    run(sys.argv[1:])
